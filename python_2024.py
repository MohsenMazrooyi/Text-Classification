from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from time import sleep
from selenium.common.exceptions import NoSuchElementException
import openpyxl
from selenium import webdriver
import re
import os

# Define the path for the Excel file
file_path = "sample_1.xlsx"

# Check if the file exists and delete it if it does
if os.path.exists(file_path):
    os.remove(file_path)
# Initialize Excel workbook
wb = openpyxl.Workbook()
sheet = wb.active

# Initialize Chrome WebDriver
driver = webdriver.Chrome()
driver.get('https://www.zoho.com/us/inventory/login.html')
driver.find_element(By.CLASS_NAME, 'zlogin-apps').click()
sleep(2)

# Login
username_field = driver.find_element(By.ID, 'login_id')
username_field.send_keys('h.wood@megabiz.com.tr')
driver.find_element(By.ID, 'nextbtn').click()
sleep(2)

password_field = driver.find_element(By.ID, 'password')
password_field.send_keys('HelenWood.Megabiz')
driver.find_element(By.ID, 'nextbtn').click()
sleep(2)

# Define headers for Excel file
headers = ['BILL DATE', 'BILL#', 'ALLOCATED TO','VENDOR NAME', 'DESCRIPTION', 'ALLOCATED AMOUNT']

for col_num, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col_num).value = header

# Initialize variables
loop_index = 2  # Starting row index for writing data

# Scraping loop
for page in range(1, 12):
    driver.get(f"https://inventory.zoho.com/app/734780606#/reports/landedcostsummary?filter_by=ThisYear&from_date=2024-01-01&page={page}&to_date=2024-12-31")
    sleep(2)

    # Find all rows in the table
    rows = driver.find_elements(By.CSS_SELECTOR, '#vertically-scrolled-table tbody tr')

    for row in rows:
        try:
            # Extract data from each row
            bill_date = row.find_element(By.CSS_SELECTOR, 'td:nth-child(1)').text
            bill_number = row.find_element(By.CSS_SELECTOR, 'td:nth-child(2) a').text
            vendor_name = row.find_element(By.CSS_SELECTOR, 'td:nth-child(3)').text
            description = row.find_element(By.CSS_SELECTOR, 'td:nth-child(4)').text
            allocated_amount_text = row.find_element(By.CSS_SELECTOR, 'td:nth-child(6)').text
            # Extract values
            allocated_amounts = re.findall(r'\$[\d,]+(?:\.\d{2})?', allocated_amount_text)

            # Extract codes (using 'for' as a delimiter)
            allocated_toes = re.findall(r'for ([\w\-/.]+)', allocated_amount_text)
            if len(allocated_amounts) <= 2:
                c1 = sheet.cell(row=loop_index, column=1)
                c1.value = bill_date

                c2 = sheet.cell(row=loop_index, column=2)
                c2.value = bill_number

                if not allocated_toes:
                    # Set the cell to None or a default value if the list is empty
                    c3 = sheet.cell(row=loop_index, column=6)
                    c3.value = None  # or use "" or 0 depending on your needs
                else:
                    c3 = sheet.cell(row=loop_index, column=6)
                    c3.value = allocated_toes[0]

                c4 = sheet.cell(row=loop_index, column=3)
                c4.value = vendor_name

                c5 = sheet.cell(row=loop_index, column=4)
                c5.value = description

                if len(allocated_amounts) > 0:
                    c6 = sheet.cell(row=loop_index, column=5)
                    c6.value = allocated_amounts[0]
                else:
                    print("No allocated amounts found.")
                    continue
                loop_index += 1

            else:
                for i,allocated_amount in enumerate(allocated_amounts[1:]):
                    c1 = sheet.cell(row=loop_index, column=1)
                    c1.value = bill_date

                    c2 = sheet.cell(row=loop_index, column=2)
                    c2.value = bill_number

                    c3 = sheet.cell(row=loop_index, column=6)
                    c3.value = allocated_toes[i]

                    c4 = sheet.cell(row=loop_index, column=3)
                    c4.value = vendor_name

                    c5 = sheet.cell(row=loop_index, column=4)
                    c5.value = description

                    c6 = sheet.cell(row=loop_index, column=5)
                    c6.value = allocated_amount
                    loop_index += 1

        except NoSuchElementException:
            print("Element not found. Skipping row.")

# Save workbook and close WebDriver
wb.save("sample_1.xlsx")
# driver.close()
